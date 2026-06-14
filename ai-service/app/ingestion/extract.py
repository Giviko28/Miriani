"""Extract plain text from uploaded documents by type (PDF, Word, Excel, text)."""

import io

from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook


class UnsupportedFileType(Exception):
    pass


def extract_text(file_name: str, data: bytes) -> str:
    """Return the plain text content of a document, dispatched by file extension."""
    ext = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""

    if ext == "pdf":
        return _from_pdf(data)
    if ext == "docx":
        return _from_docx(data)
    if ext == "xlsx":
        return _from_xlsx(data)
    if ext in ("txt", "md", "csv"):
        return data.decode("utf-8", errors="replace")

    raise UnsupportedFileType(f"Unsupported file type: .{ext}")


def _from_pdf(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages)


def _from_docx(data: bytes) -> str:
    doc = DocxDocument(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _from_xlsx(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)
